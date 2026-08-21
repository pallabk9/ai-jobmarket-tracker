#!/usr/bin/env python3
"""
Sector AI Exposure model builder (Phase 1: US, UK, EU, AU).

Computes, per region and dashboard sector, the Sector AI Exposure Score:

    SAES(sector) = 100 x SUM over occupations o of
                   [ employment_share(o | sector) x observed_exposure(o) ]

observed_exposure comes from the locked Anthropic Observed Exposure file
(Massenkoff & McCrory) published in the Anthropic Economic Index dataset:
  labor_market_impacts/job_exposure.csv  (occ_code = US SOC, 0..1 scale)

Occupation-by-industry employment matrices per region:
  US : BLS OEWS via api.bls.gov - SOC major group (22) x 3-digit NAICS
       (fine granularity; ~8 batched API queries, annual vintage)
  UK : ONS ad-hoc 3136 - SOC2020 4-digit x SIC section (static 2024 sheet),
       joined to the AWA uk_occupations.json raw exposure per SOC2020
       (fine granularity, internally consistent with the UK model)
  EU : Eurostat lfsq_eisn2 - ISCO-08 major x NACE section, quarterly
       (coarse granularity: 9 occupation majors)
  AU : ILOSTAT DF_EMP_TEMP_ECO_OCU_NB - ISCO-08 major x ISIC section
       (coarse granularity)

SOC -> ISCO-08 major mapping is a documented approximation (see SOC_TO_ISCO);
regions using it carry granularity="coarse" in their provenance block.

Output: data/sectors.json exposure blocks (signals are filled by
scripts/update_sectors.py). Existing signals in sectors.json are preserved.

Run quarterly via .github/workflows/build-sector-model.yml, or manually:
    python3 scripts/build_sector_model.py
"""

import csv
import io
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from update_data import _http_get, _http_post_json  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SECTORS_JSON = ROOT / "data" / "sectors.json"
CONCORDANCE = ROOT / "model" / "sector_concordance.csv"
UK_OCC_JSON = ROOT / "data" / "uk_occupations.json"

EXPOSURE_URL = ("https://huggingface.co/datasets/Anthropic/EconomicIndex/"
                "resolve/main/labor_market_impacts/job_exposure.csv")
ONS_3136_URL = ("https://www.ons.gov.uk/file?uri=/employmentandlabourmarket/"
                "peopleinwork/employmentandemployeetypes/adhocs/"
                "3136employmentbyoccupationandindustrysectionasuk2021to2024/"
                "employmentby4digitsoc1digitindustryatosuk20212024final.xlsx")

# SOC major group -> ISCO-08 major group. A documented approximation used
# only for the coarse-matrix regions (EU, AU): ISCO 3 (associate
# professionals) has no clean SOC major counterpart, so SOC 13 (business &
# financial operations) stands in for it. Flagged via granularity="coarse".
SOC_TO_ISCO = {
    "11": "1",
    "15": "2", "17": "2", "19": "2", "21": "2", "23": "2",
    "25": "2", "27": "2", "29": "2",
    "13": "3",
    "43": "4",
    "31": "5", "33": "5", "35": "5", "39": "5", "41": "5",
    "45": "6",
    "47": "7", "49": "7",
    "51": "8", "53": "8",
    "37": "9",
}

SOC_MAJORS = ["11", "13", "15", "17", "19", "21", "23", "25", "27", "29",
              "31", "33", "35", "37", "39", "41", "43", "45", "47", "49",
              "51", "53"]

# ISIC/NACE/SIC section letter per sector (from the concordance; section-level
# matrices share a section across sectors - e.g. banking & insurance both K).
SECTION_SECTORS = {}   # filled from concordance

def load_concordance():
    rows = list(csv.DictReader(CONCORDANCE.open(encoding="utf-8")))
    for r in rows:
        SECTION_SECTORS[r["sector_id"]] = r["isic_section"]
    return rows

# ------------------------------------------------------------------
# Exposure vectors from the locked Anthropic file
# ------------------------------------------------------------------

def fetch_exposure():
    """{soc6: exposure} + per-SOC-major means + per-ISCO-major means."""
    text = _http_get(EXPOSURE_URL, timeout=60)
    by_occ, by_major = {}, {}
    for row in csv.DictReader(io.StringIO(text)):
        code = row["occ_code"].strip()
        try:
            v = float(row["observed_exposure"])
        except (TypeError, ValueError):
            continue
        by_occ[code] = v
        by_major.setdefault(code[:2], []).append(v)
    major_mean = {m: sum(vs) / len(vs) for m, vs in by_major.items()}
    isco_vals = {}
    for soc, isco in SOC_TO_ISCO.items():
        if soc in major_mean:
            isco_vals.setdefault(isco, []).append(major_mean[soc])
    isco_mean = {i: sum(vs) / len(vs) for i, vs in isco_vals.items()}
    if len(by_occ) < 500 or len(isco_mean) < 8:
        raise ValueError(f"exposure file looks wrong: {len(by_occ)} occs, "
                         f"{len(isco_mean)} ISCO majors")
    return by_occ, major_mean, isco_mean

# ------------------------------------------------------------------
# US matrix: OEWS via api.bls.gov (SOC major x 3-digit NAICS groups)
# ------------------------------------------------------------------

def _oews_series_id(industry6, soc_major):
    # OEWS id: OE + U(seasonal) + N(national areatype) + 7-digit area
    # + 6-digit industry + 6-digit occupation + 2-digit datatype (01 =
    # employment). The area code is SEVEN zeros - a 6-zero prefix makes
    # every id 24 chars and the API silently returns no data
    # (bug found 2026-08-20 via the quarterly CI run's "too sparse: []").
    return f"OEUN0000000{industry6}{soc_major}000001"

def fetch_us_matrix(concordance):
    """{sector_id: {soc_major: employment}} via batched OEWS API calls."""
    wanted = {}   # series_id -> (sector_id, soc_major)
    for row in concordance:
        for ind in filter(None, row["naics_codes"].split(";")):
            for m in SOC_MAJORS:
                wanted[_oews_series_id(ind.strip(), m)] = (row["sector_id"], m)
    ids = list(wanted)
    key = os.environ.get("BLS_API_KEY")
    out = {}
    for i in range(0, len(ids), 50):
        payload = {"seriesid": ids[i:i + 50], "latest": "true"}
        if key:
            payload["registrationkey"] = key
        js = _http_post_json("https://api.bls.gov/publicAPI/v2/timeseries/data/",
                             payload, timeout=60)
        if js.get("status") != "REQUEST_SUCCEEDED":
            raise ValueError(f"OEWS API: {js.get('status')} {js.get('message')}")
        for s in js["Results"]["series"]:
            data = s.get("data") or []
            if not data:
                continue
            try:
                emp = float(data[0]["value"].replace(",", ""))
            except (TypeError, ValueError):
                continue
            sector_id, m = wanted[s["seriesID"]]
            out.setdefault(sector_id, {})
            out[sector_id][m] = out[sector_id].get(m, 0.0) + emp
    if len(out) < 8:
        raise ValueError(f"OEWS matrix too sparse: {sorted(out)}")
    return out

# ------------------------------------------------------------------
# UK matrix: ONS ad-hoc 3136 (SOC2020 4-digit x SIC section, 2024 sheet)
# ------------------------------------------------------------------

def fetch_uk_matrix():
    """{section_letter: {soc4: employment}} from the 3136 workbook."""
    import openpyxl  # lazy: only the model build needs it
    raw = None
    import urllib.request
    from update_data import UA
    req = urllib.request.Request(ONS_3136_URL, headers=UA)
    with urllib.request.urlopen(req, timeout=90) as r:
        raw = r.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    sheet = sorted(wb.sheetnames)[-1]          # latest year
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header_i = next(i for i, r in enumerate(rows)
                    if r[2] and str(r[2]).strip().startswith("1 A"))
    sections = {}
    for ci, cell in enumerate(rows[header_i]):
        if ci < 2 or not cell:
            continue
        parts = str(cell).split()
        if len(parts) >= 2 and len(parts[1]) == 1 and parts[1].isalpha():
            sections[ci] = parts[1].upper()
    out = {}
    for r in rows[header_i + 1:]:
        label = str(r[1] or "").strip()
        if not (len(label) > 4 and label[:4].isdigit()):
            continue
        soc4 = label[:4]
        for ci, sec in sections.items():
            v = r[ci]
            try:
                emp = float(str(v).replace(",", ""))
            except (TypeError, ValueError):
                continue
            out.setdefault(sec, {})[soc4] = out.setdefault(sec, {}).get(soc4, 0) + emp
    if "K" not in out or "C" not in out:
        raise ValueError(f"ONS 3136 parse failed: sections {sorted(out)}")
    return out, sheet

def uk_exposure_by_soc4():
    """{soc2020_4digit: raw exposure} from the AWA UK occupation model."""
    js = json.loads(UK_OCC_JSON.read_text(encoding="utf-8"))
    return {o["soc"]: float(o["raw"]) for o in js["occupations"]
            if o.get("raw") is not None}

# ------------------------------------------------------------------
# EU matrix: Eurostat lfsq_eisn2 (ISCO major x NACE section)
# ------------------------------------------------------------------

def _jsonstat_cells(js, want_dims):
    """Decode a JSON-stat 2.0 response into {(dimval...): value} for the
    requested dimension ids, using the id/size arrays for flat-index math."""
    ids, sizes = js["id"], js["size"]
    strides, s = [0] * len(ids), 1
    for i in range(len(ids) - 1, -1, -1):
        strides[i] = s
        s *= max(sizes[i], 1)
    catidx = {d: js["dimension"][d]["category"]["index"] for d in ids}
    inv = {d: {v: k for k, v in catidx[d].items()} for d in ids}
    out = {}
    for key, val in js["value"].items():
        flat = int(key)
        coord = []
        for i, d in enumerate(ids):
            di = (flat // strides[i]) % max(sizes[i], 1)
            coord.append(inv[d].get(di))
        m = dict(zip(ids, coord))
        out[tuple(m.get(d) for d in want_dims)] = float(val)
    return out

def fetch_eu_matrix():
    """{section_letter: {isco_major: employment_thousands}} - the EU27
    aggregate lags, so pull the last 4 quarters and keep the newest
    populated one per section."""
    # expand compound sectors ('D+E') into their component sections
    sections = sorted({p for s in SECTION_SECTORS.values()
                       for p in s.split("+")})
    out = {}
    for sec in sections:
        url = ("https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/"
               "data/lfsq_eisn2?format=JSON&lang=EN&geo=EU27_2020&sex=T"
               "&age=Y_GE15&unit=THS_PER&lastTimePeriod=4"
               f"&nace_r2={sec}"
               + "".join(f"&isco08=OC{i}" for i in range(1, 10)))
        js = json.loads(_http_get(url, timeout=60))
        cells = _jsonstat_cells(js, ("time", "isco08"))
        by_time = {}
        for (t, oc), v in cells.items():
            if oc:
                by_time.setdefault(t, {})[oc[-1]] = v
        if by_time:
            latest = max(t for t, d in by_time.items() if d)
            out[sec] = by_time[latest]
    if len(out) < 6:
        raise ValueError(f"Eurostat eisn2 too sparse: {sorted(out)}")
    return out

# ------------------------------------------------------------------
# AU / IN / APAC matrices: ILOSTAT ECO x OCU (ISCO major x ISIC section)
# ------------------------------------------------------------------

def fetch_ilo_matrix(areas):
    """{section_letter: {isco_major: employment}} from ILOSTAT ECO x OCU,
    ANNUAL frequency - the quarterly series carry aggregate categories only
    for most countries, while the annual series has full ISIC sections
    (verified: IND/SGP/KOR 2025, JPN 2023, AUS 2025). Multiple areas are
    POOLED by summing employment per cell - used for the APAC composite
    (SGP+JPN+KOR). Returns (matrix, periods) with periods per area."""
    key = "+".join(areas)
    text = _http_get("https://sdmx.ilo.org/rest/data/"
                     f"ILO,DF_EMP_TEMP_ECO_OCU_NB,1.0/{key}.A...."
                     "?lastNObservations=1&format=csv", timeout=120)
    out, periods = {}, {}
    for row in csv.DictReader(io.StringIO(text)):
        eco, ocu = row.get("ECO", ""), row.get("OCU", "")
        if not eco.startswith("ECO_ISIC4_") or len(eco) != len("ECO_ISIC4_A"):
            continue
        if not ocu.startswith("OCU_ISCO08_") or ocu[-1] not in "123456789":
            continue
        try:
            v = float(row["OBS_VALUE"])
        except (TypeError, ValueError):
            continue
        out.setdefault(eco[-1], {})
        out[eco[-1]][ocu[-1]] = out[eco[-1]].get(ocu[-1], 0.0) + v
        periods[row.get("REF_AREA", "?")] = row.get("TIME_PERIOD", "?")
    if len(out) < 10:
        raise ValueError(f"ILO {key} matrix too sparse: {sorted(out)}")
    return out, periods

# ------------------------------------------------------------------
# Scoring
# ------------------------------------------------------------------

def saes(matrix_for_sector, exposure_by_key):
    """Weighted-mean exposure x100 + top contributors, from
    {occ_key: employment} and {occ_key: exposure}."""
    pairs = [(k, emp, exposure_by_key[k]) for k, emp in matrix_for_sector.items()
             if k in exposure_by_key and emp > 0]
    total = sum(emp for _, emp, _ in pairs)
    if not pairs or total <= 0:
        return None, []
    score = 100.0 * sum(emp * ex for _, emp, ex in pairs) / total
    contrib = sorted(pairs, key=lambda p: p[1] * p[2], reverse=True)[:5]
    top = [{"key": k, "share": round(emp / total, 4),
            "exposure": round(ex * 100, 2)} for k, emp, ex in contrib]
    return round(score, 2), top

SOC_MAJOR_LABEL = {
    "11": "Management", "13": "Business & financial operations",
    "15": "Computer & mathematical", "17": "Architecture & engineering",
    "19": "Life/physical/social science", "21": "Community & social service",
    "23": "Legal", "25": "Education & library", "27": "Arts/design/media",
    "29": "Healthcare practitioners", "31": "Healthcare support",
    "33": "Protective service", "35": "Food preparation & serving",
    "37": "Building & grounds cleaning", "39": "Personal care & service",
    "41": "Sales", "43": "Office & administrative support",
    "45": "Farming/fishing/forestry", "47": "Construction & extraction",
    "49": "Installation/maintenance/repair", "51": "Production",
    "53": "Transportation & material moving",
}
ISCO_LABEL = {
    "1": "Managers", "2": "Professionals",
    "3": "Technicians & associate professionals", "4": "Clerical support",
    "5": "Service & sales workers", "6": "Skilled agricultural",
    "7": "Craft & related trades", "8": "Plant & machine operators",
    "9": "Elementary occupations",
}

def label_contribs(top, labels):
    for t in top:
        t["label"] = labels.get(t.pop("key"), "Other")
    return top

def main():
    concordance = load_concordance()
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    by_occ, soc_major_mean, isco_mean = fetch_exposure()
    print(f"exposure: {len(by_occ)} occupations loaded")

    # preserve existing signals if sectors.json already exists
    existing = {}
    if SECTORS_JSON.exists():
        existing = json.loads(SECTORS_JSON.read_text(encoding="utf-8"))

    regions_out = existing.get("regions", {})

    def build_region(code, matrix_by_sector_or_section, exposure_vec, labels,
                     provenance, section_level):
        sectors = regions_out.setdefault(code, {}).setdefault("sectors", {})
        scored = []
        for row in concordance:
            sid = row["sector_id"]
            if section_level:
                # 'D+E' marks a sector spanning two sections: merge their
                # occupation employment before scoring
                parts = row["isic_section"].split("+")
                if len(parts) == 1:
                    mat = matrix_by_sector_or_section.get(parts[0], {})
                else:
                    mat = {}
                    for sc in parts:
                        for occ, emp in (matrix_by_sector_or_section.get(sc)
                                         or {}).items():
                            mat[occ] = mat.get(occ, 0.0) + emp
            else:
                mat = matrix_by_sector_or_section.get(sid, {})
            score, top = saes(mat, exposure_vec)
            node = sectors.setdefault(sid, {})
            node["label"] = row["label"]
            node["exposure"] = score
            node["top_occupations"] = label_contribs(top, labels)
            node["shared_section"] = bool(section_level) and \
                sum(1 for r in concordance
                    if r["isic_section"] == row["isic_section"]) > 1
            scored.append((sid, score if score is not None else -1))
        for rank, (sid, _) in enumerate(
                sorted(scored, key=lambda t: t[1], reverse=True), 1):
            sectors[sid]["exposure_rank"] = rank
        # Within-region relative index (top sector = 100). Raw SAES scales
        # differ by exposure basis (Anthropic observed vs AWA raw for UK),
        # and cross-region comparison is disallowed anyway - the relative
        # index is what the heat strip and heatmap display.
        top = max((s for _, s in scored if s is not None and s > 0), default=None)
        for sid, s in scored:
            sectors[sid]["exposure_rel"] = (
                round(100.0 * s / top, 1) if (top and s and s > 0) else None)
        regions_out[code]["matrix"] = provenance

    # US ---------------------------------------------------------------
    try:
        us = fetch_us_matrix(concordance)
        build_region("US", us, soc_major_mean, SOC_MAJOR_LABEL, {
            "source": "BLS OEWS (latest vintage, SOC major x 3-digit NAICS, api.bls.gov)",
            "url": "https://www.bls.gov/oes/", "granularity": "fine",
            "built_at": stamp}, section_level=False)
        print("US matrix ok:", {k: len(v) for k, v in list(us.items())[:3]})
    except Exception as exc:  # noqa: BLE001 - keep other regions alive
        print(f"US matrix FAILED, keeping previous block: {exc}")

    # UK ---------------------------------------------------------------
    try:
        uk_matrix, uk_sheet = fetch_uk_matrix()
        uk_expo = uk_exposure_by_soc4()
        uk_titles = {o["soc"]: o.get("title") or o["soc"] for o in
                     json.loads(UK_OCC_JSON.read_text())["occupations"]}
        build_region("UK", uk_matrix, uk_expo, uk_titles, {
            "source": f"ONS ad-hoc 3136 APS {uk_sheet} (SOC2020 4-digit x SIC section)"
                      " x AWA UK occupation model raw exposure",
            "url": ONS_3136_URL, "granularity": "fine",
            "built_at": stamp}, section_level=True)
        print("UK matrix ok: sections", sorted(uk_matrix))
    except Exception as exc:  # noqa: BLE001
        print(f"UK matrix FAILED, keeping previous block: {exc}")

    # EU ---------------------------------------------------------------
    try:
        eu = fetch_eu_matrix()
        build_region("EU", eu, isco_mean, ISCO_LABEL, {
            "source": "Eurostat lfsq_eisn2 (ISCO-08 major x NACE section, EU27, latest quarter)"
                      " - SOC->ISCO major approximation",
            "url": "https://ec.europa.eu/eurostat/databrowser/view/lfsq_eisn2/default/table",
            "granularity": "coarse", "built_at": stamp}, section_level=True)
        print("EU matrix ok: sections", sorted(eu))
    except Exception as exc:  # noqa: BLE001
        print(f"EU matrix FAILED, keeping previous block: {exc}")

    # AU / IN / APAC via the same ILO flow ------------------------------
    for code, areas in (("AU", ["AUS"]), ("IN", ["IND"]),
                        ("APAC", ["SGP", "JPN", "KOR"])):
        try:
            mat, periods = fetch_ilo_matrix(areas)
            per_str = ", ".join(f"{a} {periods.get(a, '?')}" for a in areas)
            build_region(code, mat, isco_mean, ISCO_LABEL, {
                "source": f"ILOSTAT EMP_TEMP_ECO_OCU (ISCO-08 major x ISIC section, {per_str})"
                          " - SOC->ISCO major approximation"
                          + (" - APAC composite pools the three markets"
                             if len(areas) > 1 else ""),
                "url": "https://ilostat.ilo.org/", "granularity": "coarse",
                "built_at": stamp}, section_level=True)
            print(f"{code} matrix ok: sections", "".join(sorted(mat)))
        except Exception as exc:  # noqa: BLE001
            print(f"{code} matrix FAILED, keeping previous block: {exc}")

    out = {
        "model_built_at": stamp,
        "methodology": "Sector AI Exposure Score = 100 x employment-share-weighted mean "
                       "of Anthropic Observed Exposure across the sector's occupation mix. "
                       "Within-region ranks only; not comparable across regions.",
        "exposure_source": {
            "name": "Anthropic Economic Index - labor_market_impacts/job_exposure.csv "
                    "(Massenkoff & McCrory)",
            "url": "https://huggingface.co/datasets/Anthropic/EconomicIndex",
        },
        "taxonomy": [{"id": r["sector_id"], "label": r["label"],
                      "section": r["isic_section"]} for r in concordance],
        "generated_at": existing.get("generated_at"),
        "regions": regions_out,
    }
    SECTORS_JSON.write_text(json.dumps(out, indent=1), encoding="utf-8")
    print(f"Wrote {SECTORS_JSON} ({len(json.dumps(out)) // 1024}KB)")
    return 0

if __name__ == "__main__":
    sys.exit(main())
