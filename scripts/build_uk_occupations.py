#!/usr/bin/env python3
"""
Build UK occupation exposure data for the dashboard from the AWA AI Impact
Assessment model (SOC 2020, 412 unit groups, task-decomposition).

Outputs:
  data/uk_occupations.json   - 26 sub-major groups + 412 occupations w/ task detail
and patches the UK gap_chart + capability_gap KPI inside data/current.json
(and the latest weekly snapshot).

Update cadence:
  * Quarterly  - re-export the model with recalibrated frontier task scores,
                 drop it at MODEL_PATH, rerun this script.
  * Annually   - refresh ONS SOC 2020 employment weights in the model, rerun.
Run:  python3 scripts/build_uk_occupations.py [path-to-model.xlsx]
"""
import json, sys, datetime
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DEFAULT_MODEL = ROOT / "model" / "AWA_AI_412_Occupations.xlsx"
MODEL_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_MODEL
AS_OF = "May 2026 frontier"
SRC_NAME = "AWA AI Impact Assessment (SOC 2020 task-decomposition, " + AS_OF + ")"
SRC_URL  = "https://ai-jobmarket-tracker.netlify.app/#uk-occupations"

TASK_LABELS = {
 "T01":"Structured Data Entry & Processing","T02":"Routine Text Generation",
 "T03":"Information Retrieval & Synthesis","T04":"Scheduling & Coordination",
 "T05":"Routine Numerical Analysis","T06":"Classification & Sorting",
 "T07":"Complex Document Drafting","T08":"Customer/Client Communication",
 "T09":"Translation & Language Services","T10":"Technical Analysis & Problem-Solving",
 "T11":"Creative Design & Content","T12":"Strategic Judgement & Decision-Making",
 "T13":"Negotiation & Persuasion","T14":"Teaching, Training & Mentoring",
 "T15":"Supervision & People Management","T16":"Physical Care & Personal Service",
 "T17":"Skilled Manual & Technical Operations","T18":"Outdoor/Physical Labour",
}

def num(v):
    return v if isinstance(v,(int,float)) else None

def main():
    wb = openpyxl.load_workbook(MODEL_PATH, data_only=True)

    # --- sub-major groups (gap chart source) ---
    groups=[]
    for r in wb["Sub-Major Group Summary"].iter_rows(min_row=2, values_only=True):
        code=r[0]
        if code in (None,"TOTALS"): continue
        groups.append({"code":str(code),"name":r[1],"count":r[2],
                       "employment_000":r[3],"raw":r[4],"practical":r[5],"hrs":r[6]})

    # --- per-occupation core metrics (Calculation) ---
    occ={}
    for r in wb["Calculation"].iter_rows(min_row=2, values_only=True):
        soc=r[0]
        if soc is None: continue
        soc=str(soc).strip()
        if not (soc.isdigit() and len(soc)==4): continue   # skip TOTALS / header rows
        occ[soc]={"soc":soc,"title":r[1],"smg_code":soc[:2],"smg":r[2],
                  "employment":num(r[3]),"raw":num(r[4]),"pi":num(r[5]),
                  "hrs_week":num(r[6]),"annual_hrs":num(r[7]),"fte":num(r[9])}

    # --- amplification + combined hrs (May 2026 Frontier) ---
    for r in wb["May 2026 Frontier"].iter_rows(min_row=4, values_only=True):
        soc=r[0]
        if soc is None: continue
        soc=str(soc)
        if soc in occ:
            occ[soc]["discount"]=num(r[4]); occ[soc]["mult"]=num(r[8])
            occ[soc]["combined_hrs"]=num(r[10])

    # --- 18 task-time allocations (Role Profiles) ---
    rp=wb["Role Profiles"]; hdr=[c.value for c in rp[1]]
    tcols={hdr[i].split("\n")[0]:i for i in range(3,21)}   # T01..T18 -> col idx
    for r in rp.iter_rows(min_row=2, values_only=True):
        soc=r[0]
        if soc is None: continue
        soc=str(soc)
        if soc in occ:
            occ[soc]["tasks"]={t:num(r[i]) for t,i in tcols.items()}

    occ_list=sorted(occ.values(), key=lambda o:o["soc"])
    out={"as_of":AS_OF,"generated_at":datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
         "source":SRC_NAME,"task_labels":TASK_LABELS,
         "total_employment":34250030,"n_occupations":len(occ_list),
         "groups":groups,"occupations":occ_list}
    (DATA/"uk_occupations.json").write_text(json.dumps(out,separators=(",",":")))
    print(f"Wrote uk_occupations.json: {len(groups)} groups, {len(occ_list)} occupations")

    # headline capability gap = mean(raw) - mean(practical), pp
    mr=sum(g["raw"]*g["count"] for g in groups)/sum(g["count"] for g in groups)
    mp=sum(g["practical"]*g["count"] for g in groups)/sum(g["count"] for g in groups)
    gap_pp=round((mr-mp)*100,2)

    gap_chart={"cats":[g["code"] for g in groups],
               "names":[g["name"] for g in groups],
               "theoretical":[round(g["raw"]*100,1) for g in groups],
               "observed":[round(g["practical"]*100,1) for g in groups],
               "source":SRC_NAME,"detail":"uk_occupations.json"}

    for fn in ["current.json", f"snapshots/{json.loads((DATA/'current.json').read_text())['iso_week']}.json"]:
        p=DATA/fn
        if not p.exists(): continue
        d=json.loads(p.read_text())
        d["regions"]["UK"]["gap_chart"]=gap_chart
        k=d["regions"]["UK"]["kpis"]["capability_gap"]
        k["value"]=gap_pp; k["source"]=SRC_NAME; k["source_url"]=SRC_URL; k["measurement"]="modelled"
        p.write_text(json.dumps(d,indent=2))
        print(f"Patched {fn}: UK gap_chart ({len(groups)} groups) + capability_gap={gap_pp}pp")

if __name__=="__main__":
    sys.exit(main())
