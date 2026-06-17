#!/usr/bin/env python3
"""
US occupation AI-impact build from O*NET + BLS OES.

Pulls O*NET Generalized Work Activities (importance scores) and BLS OES
employment, applies the GWA->18-task crosswalk, and produces the same
data shape the dashboard already reads for the UK:
  data/us_occupations.json   (groups + occupations w/ task detail)
and patches the US gap_chart + capability_gap inside data/current.json.

Why this is the easy region: O*NET publishes task-level data, so the
per-occupation task-time allocation (the hand-estimated step in the UK
model) is derived directly instead of by hand.

METHOD
  task weight(occupation, task) = SUM over GWAs mapped to that task of
      (IM - 1)^2 * crosswalk_weight,           IM = O*NET importance (1-5)
  normalised across the 18 tasks to sum to 100% (a time-allocation proxy).
  raw exposure   = SUM(task% * susceptibility)          [task_scores.csv]
  practical      = raw * adoption_discount              [soc_major_groups.csv]
  hours freed    = practical * 40h week ; FTE = hours*emp/40
  combined hours = hours * amplification_multiplier
The (IM-1)^2 relevance transform concentrates weight on core activities;
without it O*NET importance is too flat and compresses the spread.

NOTE ON COMPARABILITY: US task weights come from O*NET importance; the UK
model uses hand-estimated time shares. Rankings and within-region gaps are
comparable; absolute US-vs-UK levels are not. Flagged in the dashboard.

NETWORK: O*NET and BLS must be reachable. Runs from a normal machine or CI
(GitHub Actions). A firewalled sandbox cannot reach them. Sources can be
overridden with local files via env vars (see CONFIG) for offline runs.

Usage:  python3 build_us_from_onet.py
"""
import csv, io, json, os, sys, zipfile, datetime, urllib.request
from pathlib import Path

HERE = Path(__file__).resolve().parent
RB   = HERE.parent                      # model/regional_build
SITE = RB.parent.parent                 # netlify-site
DATA = SITE / "data"
CACHE = HERE / "data_cache"; CACHE.mkdir(exist_ok=True)

ONET_VER = os.environ.get("ONET_VER", "db_30_3")
ONET_BASE = f"https://www.onetcenter.org/dl_files/database/{ONET_VER}_text"
WA_URL  = os.environ.get("ONET_WA_URL",  f"{ONET_BASE}/Work%20Activities.txt")
OCC_URL = os.environ.get("ONET_OCC_URL", f"{ONET_BASE}/Occupation%20Data.txt")
OES_ZIP = os.environ.get("BLS_OES_ZIP",  "https://www.bls.gov/oes/special-requests/oesm23nat.zip")
# local overrides (offline / testing)
WA_FILE  = os.environ.get("ONET_WA_FILE")
OCC_FILE = os.environ.get("ONET_OCC_FILE")
OES_XLSX = os.environ.get("BLS_OES_XLSX")     # already-extracted national_*_dl.xlsx
UA = {"User-Agent": "ai-jobmarket-tracker/1.0 (research; pallabk9 github)"}
WEEK_HOURS = 40.0; WORKING_WEEKS = 48.0       # US full-time basis

def fetch_text(url, cache_name):
    cp = CACHE / cache_name
    if cp.exists():
        return cp.read_text(encoding="utf-8", errors="replace")
    print(f"  downloading {url}")
    req = urllib.request.Request(url, headers=UA)
    body = urllib.request.urlopen(req, timeout=90).read().decode("utf-8", errors="replace")
    cp.write_text(body, encoding="utf-8")
    return body

def load_scores():
    s, label = {}, {}
    for r in csv.DictReader((RB / "task_scores.csv").open()):
        s[r["task_id"]] = float(r["ai_susceptibility_may2026"]); label[r["task_id"]] = r["task_category"]
    return s, label

def load_crosswalk():
    cx = {}
    for r in csv.DictReader((HERE / "gwa_task_crosswalk.csv").open()):
        cx.setdefault(r["element_id"], []).append((r["task_id"], float(r["weight"])))
    return cx

def load_soc_groups():
    g = {}
    for r in csv.DictReader((HERE / "soc_major_groups.csv").open()):
        g[r["group_code"]] = (r["group_name"], float(r["adoption_discount"]), float(r["amplification_multiplier"]))
    return g

def parse_onet_importance():
    """{onetsoc: {element_id: IM}} from Work Activities.txt (IM scale only)."""
    text = Path(WA_FILE).read_text(encoding="utf-8", errors="replace") if WA_FILE else fetch_text(WA_URL, "work_activities.txt")
    rows = list(csv.DictReader(io.StringIO(text), delimiter="\t"))
    out = {}
    for r in rows:
        if r.get("Scale ID") != "IM":
            continue
        try:
            out.setdefault(r["O*NET-SOC Code"], {})[r["Element ID"]] = float(r["Data Value"])
        except (TypeError, ValueError):
            continue
    if not out:
        raise SystemExit("No O*NET importance rows parsed - check Work Activities source")
    return out

def parse_onet_titles():
    text = Path(OCC_FILE).read_text(encoding="utf-8", errors="replace") if OCC_FILE else fetch_text(OCC_URL, "occupation_data.txt")
    rows = csv.DictReader(io.StringIO(text), delimiter="\t")
    return {r["O*NET-SOC Code"]: r["Title"] for r in rows}

def parse_bls_oes():
    """{soc6: (title, employment)} for detailed occupations from the OES national xlsx."""
    import openpyxl
    if OES_XLSX:
        wb = openpyxl.load_workbook(OES_XLSX, read_only=True, data_only=True)
    else:
        zp = CACHE / "oes_nat.zip"
        if not zp.exists():
            print(f"  downloading {OES_ZIP}")
            req = urllib.request.Request(OES_ZIP, headers=UA)
            zp.write_bytes(urllib.request.urlopen(req, timeout=120).read())
        zf = zipfile.ZipFile(zp)
        name = next(n for n in zf.namelist() if n.lower().endswith(".xlsx"))
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)), read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    cc, ct, ce = ix.get("OCC_CODE"), ix.get("OCC_TITLE"), ix.get("TOT_EMP")
    cg = ix.get("O_GROUP", ix.get("OCC_GROUP"))
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if cg is not None and str(row[cg]).strip().lower() not in ("detailed", "detail"):
            continue
        code = str(row[cc]).strip()
        emp = row[ce]
        try:
            emp = float(str(emp).replace(",", ""))
        except (TypeError, ValueError):
            emp = 0.0
        out[code] = (row[ct], emp)
    if not out:
        raise SystemExit("No OES detailed rows parsed - check BLS OES source")
    return out

def main():
    sus, label = load_scores()
    cx = load_crosswalk()
    groups = load_soc_groups()
    TASKS = [f"T{i:02d}" for i in range(1, 19)]

    print("Reading O*NET ...")
    imp = parse_onet_importance()
    titles = parse_onet_titles()
    print(f"  {len(imp)} O*NET-SOC occupations")
    print("Reading BLS OES ...")
    oes = parse_bls_oes()
    print(f"  {len(oes)} detailed SOC employment rows")

    # O*NET importance -> task allocation per O*NET-SOC, then aggregate to SOC6
    soc_alloc = {}   # soc6 -> list of per-onetsoc 18-vectors
    soc_title = {}
    for onetsoc, ims in imp.items():
        tw = {t: 0.0 for t in TASKS}
        for eid, im in ims.items():
            for task, w in cx.get(eid, []):
                tw[task] += ((max(0.0, im - 1.0)) ** 2) * w
        tot = sum(tw.values())
        if tot <= 0:
            continue
        vec = {t: tw[t] / tot * 100.0 for t in TASKS}
        soc6 = onetsoc.split(".")[0]
        soc_alloc.setdefault(soc6, []).append(vec)
        soc_title.setdefault(soc6, titles.get(onetsoc, onetsoc))

    rows = []
    for soc6, vecs in soc_alloc.items():
        gc = soc6[:2]
        if gc not in groups:
            continue
        gname, disc, mult = groups[gc]
        alloc = {t: round(sum(v[t] for v in vecs) / len(vecs)) for t in TASKS}
        # fix rounding drift to exactly 100
        drift = 100 - sum(alloc.values())
        if drift:
            tmax = max(TASKS, key=lambda t: alloc[t]); alloc[tmax] += drift
        title, emp = oes.get(soc6, (soc_title.get(soc6, soc6), 0.0))
        rows.append({"occ_code": soc6, "occ_title": title or soc_title.get(soc6, soc6),
                     "group_code": gc, "group_name": gname, "employment": int(emp),
                     "adoption_discount": disc, "amplification_multiplier": mult, **alloc})

    rows.sort(key=lambda r: r["occ_code"])
    cols = ["occ_code", "occ_title", "group_code", "group_name", "employment",
            "adoption_discount", "amplification_multiplier"] + TASKS
    csv_path = RB / "regions" / "US_occupations.csv"
    with csv_path.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols); w.writeheader(); w.writerows(rows)
    print(f"Wrote {csv_path.name}: {len(rows)} SOC occupations")

    # ---- compute the dashboard JSON (same engine math as build_region) ----
    occ = []
    for r in rows:
        raw = sum(r[t] * sus[t] for t in TASKS) / 100.0
        pi = raw * r["adoption_discount"]; hrs = pi * WEEK_HOURS
        occ.append({"soc": r["occ_code"], "title": r["occ_title"], "smg_code": r["group_code"], "smg": r["group_name"],
                    "employment": r["employment"], "discount": round(r["adoption_discount"], 3),
                    "raw": round(raw, 4), "pi": round(pi, 4), "hrs_week": round(hrs, 1),
                    "annual_hrs": round(hrs * WORKING_WEEKS), "fte": round(hrs * r["employment"] / WEEK_HOURS),
                    "mult": r["amplification_multiplier"], "combined_hrs": round(hrs * r["amplification_multiplier"], 2),
                    "tasks": {t: r[t] for t in TASKS}})

    grp = {}
    for o in occ:
        g = grp.setdefault(o["smg_code"], {"code": o["smg_code"], "name": o["smg"], "count": 0, "emp": 0.0, "rw": 0.0, "pw": 0.0})
        g["count"] += 1; g["emp"] += o["employment"]; g["rw"] += o["raw"] * o["employment"]; g["pw"] += o["pi"] * o["employment"]
    glist = []
    for g in sorted(grp.values(), key=lambda x: x["code"]):
        e = g["emp"] or 1
        glist.append({"code": g["code"], "name": g["name"], "count": g["count"],
                      "employment_000": round(g["emp"] / 1000),
                      "raw": round(g["rw"] / e, 4), "practical": round(g["pw"] / e, 4),
                      "hrs": round(g["pw"] / e * WEEK_HOURS, 1)})

    out = {"region": "US", "as_of": "May 2026 frontier (O*NET " + ONET_VER + ")",
           "generated_at": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
           "source": "AWA AI Impact model (O*NET GWA importance x 18-task crosswalk; BLS OES employment)",
           "task_labels": label, "total_employment": round(sum(o["employment"] for o in occ)),
           "n_occupations": len(occ), "groups": glist, "occupations": occ}
    (DATA / "us_occupations.json").write_text(json.dumps(out, separators=(",", ":")))
    print(f"Wrote us_occupations.json: {len(glist)} groups, {len(occ)} occupations")

    # ---- patch current.json US gap_chart + capability_gap ----
    mr = sum(g["raw"] * g["count"] for g in glist) / sum(g["count"] for g in glist)
    mp = sum(g["practical"] * g["count"] for g in glist) / sum(g["count"] for g in glist)
    gap_pp = round((mr - mp) * 100, 2)
    gap_chart = {"cats": [g["code"] for g in glist], "names": [g["name"] for g in glist],
                 "theoretical": [round(g["raw"] * 100, 1) for g in glist],
                 "observed": [round(g["practical"] * 100, 1) for g in glist],
                 "source": out["source"], "detail": "us_occupations.json"}
    for fn in ["current.json"]:
        p = DATA / fn
        if not p.exists():
            continue
        d = json.loads(p.read_text())
        d["regions"]["US"]["gap_chart"] = gap_chart
        k = d["regions"]["US"]["kpis"]["capability_gap"]
        k["value"] = gap_pp; k["source"] = out["source"]
        k["source_url"] = "https://www.onetcenter.org/database.html"; k["measurement"] = "modelled"
        p.write_text(json.dumps(d, indent=2))
        print(f"Patched {fn}: US gap_chart ({len(glist)} groups) + capability_gap={gap_pp}pp")

if __name__ == "__main__":
    sys.exit(main())
